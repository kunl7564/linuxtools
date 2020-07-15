import os
import xlsxwriter
from openpyxl import load_workbook, Workbook


class Write_Excel:


      def __init__(self,filename):
          self.filename=filename
          global workbook
          workbook =xlsxwriter.Workbook(filename)

      def add_sheet(self,name):
          sheet=workbook._add_sheet(self,name)

      def write_Excel(self,name,row,col,value):
          sheet=workbook._add_sheet(name)
          write_cell=sheet.write(row,col,value)
          workbook.close()
          # print(1)

class Write_excel(object):
    '''修改excel数据'''
    def __init__(self, filename,sheetname):
        self.filename = filename
        self.sheetname =sheetname
        if not os.path.exists(self.filename):
            self.wb =Workbook()
            self.ws =self.wb.create_sheet(self.sheetname,0)
        else:
            self.wb = load_workbook(self.filename)
        self.ws = self.wb.active  # 激活sheet
    def write(self, row_n, col_n, value):
        '''写入数据，如(2,3，"hello"),第二行第三列写入数据"hello"'''
        self.ws=self.wb.get_sheet_by_name(self.sheetname)
        self.ws.cell(row_n, col_n).value = value
        self.wb.save(self.filename)

